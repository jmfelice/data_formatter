import pandas as pd
import tempfile
import os
import string
from openpyxl import load_workbook
from casefy import snakecase
from .excel_attributes import get_dataframe_attributes, get_dataframe_cell_range
from .excel_dataframe_styles import format_hyperlink, column_format_standard
from .excel_dataframe_styles import column_format_header_1, alternate_color_rows
from .util import unique_string, proper_case


def create_workbook(file = None):
    if file is None:
        with tempfile.NamedTemporaryFile(delete = False, suffix = '.xlsx') as tmp:
            temp_file_path = tmp.name

    # Write the DataFrame to the temporary Excel file
    w = pd.ExcelWriter(temp_file_path, date_format = 'mmm-yyyy', datetime_format = 'mmm-yyyy')
    return w


def quick_output(df):
    # Create a temporary file
    with tempfile.NamedTemporaryFile(delete = False, suffix = '.xlsx') as tmp:
        temp_file_path = tmp.name

    # Write the DataFrame to the temporary Excel file
    w = pd.ExcelWriter(temp_file_path, date_format = 'mmm-yyyy', datetime_format = 'mmm-yyyy')
    df.to_excel(w)
    w.sheets['Sheet1'].autofit()
    w.close()

    # Open the temporary Excel file
    if os.name == 'nt':  # For Windows
        os.startfile(temp_file_path)
    elif os.name == 'posix':  # For macOS and Linux
        os.system(f"open {temp_file_path}")
    else:
        print(f"Please open the file manually: {temp_file_path}")


def add_named_region(w, sheet_name = 'Sheet1', cell_range = "$A$1:$A$1", name_of_region = None):
    if name_of_region is None:
        return None

    name_of_region = snakecase(name_of_region)

    if len(w.book.defined_names) > 0:
        list_of_named_regions = list(map(list, zip(*w.book.defined_names)))[0]
        name_of_region = unique_string(name_of_region, list_of_named_regions)
    w.book.defined_names.append([name_of_region, -1, f'{sheet_name}!{cell_range}', False])
    return name_of_region


def add_dataframe_below(w, df, sheet_name = 'Sheet1', startcol = 0, name_of_region = None):
    if sheet_name not in w.sheets:
        w.book.add_worksheet(sheet_name)

    if w.sheets[sheet_name].dim_rowmax is None:
        startrow = 0
    else:
        startrow = w.sheets[sheet_name].dim_rowmax + 3

    if name_of_region is not None:
        name_of_region = name_of_region.translate(str.maketrans('', '', string.punctuation))
        name_of_region = " ".join(name_of_region.split()).replace(" ", "_").lower()

    cell_range = get_dataframe_cell_range(df, startrow, startcol)
    name_of_region = add_named_region(w, sheet_name, cell_range, name_of_region)

    if not hasattr(w.sheets[sheet_name], 'attrs'):
        w.sheets[sheet_name].attrs = {}

    w.sheets[sheet_name].attrs[name_of_region] = get_dataframe_attributes(df, startcol, startrow)

    if isinstance(df, pd.io.formats.style.Styler):
        df.data.columns.names = [None for _ in df.columns.names]
    if isinstance(df, pd.DataFrame):
        df.columns.names = [None for _ in df.columns.names]

    df.to_excel(w, sheet_name = sheet_name, startrow = startrow, startcol = startcol)
    w.sheets[sheet_name].set_row(startrow, 20)
    w.sheets[sheet_name].set_column(startcol, startcol, 35)
    col_count = len(df.columns)
    w.sheets[sheet_name].set_column(startcol + 1, startcol + col_count, 15)


def add_dataframe_right(w, df, sheet_name = 'Sheet1', startrow = 0, name_of_region = None):
    if sheet_name not in w.sheets:
        w.book.add_worksheet(sheet_name)

    if w.sheets[sheet_name].dim_colmax is None:
        startcol = 0
    else:
        startcol = w.sheets[sheet_name].dim_colmax + 2

    if name_of_region is not None:
        name_of_region = name_of_region.translate(str.maketrans('', '', string.punctuation))
        name_of_region = " ".join(name_of_region.split()).replace(" ", "_").lower()

    cell_range = get_dataframe_cell_range(df, startrow, startcol)
    name_of_region = add_named_region(w, sheet_name, cell_range, name_of_region)

    if not hasattr(w.sheets[sheet_name], 'attrs'):
        w.sheets[sheet_name].attrs = {}

    if isinstance(df, pd.io.formats.style.Styler):
        df.data.columns.names = [None for _ in df.columns.names]
    if isinstance(df, pd.DataFrame):
        df.columns.names = [None for _ in df.columns.names]

    df.to_excel(w, sheet_name = sheet_name, startrow = startrow, startcol = startcol)

    w.sheets[sheet_name].attrs[name_of_region] = get_dataframe_attributes(df, startcol, startrow)
    w.sheets[sheet_name].set_row(startrow, 20)
    w.sheets[sheet_name].set_column(startcol, startcol, 35)
    col_count = len(df.columns)
    w.sheets[sheet_name].set_column(startcol + 1, startcol + col_count, 15)


def add_dataframes_below(w, df, sheet_name = 'Sheet1', startcol = 0, name_of_region = None):
    if isinstance(df, pd.DataFrame):
        add_dataframe_below(
            w,
            df = df,
            sheet_name = sheet_name,
            startcol = startcol,
            name_of_region = name_of_region
            )
        return None

    if isinstance(df, list):
        if not isinstance(name_of_region, list):
            print("List of region names must be the same length as the list of dataframes.")
            return None
        for i, data in enumerate(df):
            add_dataframe_below(
                w,
                df = data,
                sheet_name = sheet_name,
                startcol = startcol,
                name_of_region = name_of_region[i]
                )
        return None


def add_dataframes_right(w, df, sheet_name = 'Sheet1', startrow = 0, name_of_region = None):
    if isinstance(df, pd.DataFrame):
        add_dataframe_right(w, df, sheet_name = sheet_name, startrow = startrow)
        return None

    if isinstance(df, list):
        if not isinstance(name_of_region, list):
            print("List of region names must be the same length as the list of dataframes.")
        for i, data in enumerate(df):
            add_dataframe_right(
                w,
                df = data,
                sheet_name = sheet_name,
                startrow = startrow,
                name_of_region = name_of_region[i]
                )
        return None


def format_page(w, sheet_name):
    w.sheets[sheet_name].set_landscape()
    w.sheets[sheet_name].hide_gridlines(2)
    w.sheets[sheet_name].center_horizontally()
    w.sheets[sheet_name].set_margins(.25, .25, .75, .75)
    w.sheets[sheet_name].set_footer('&LFisher Auto Parts Financials &R&D Page: &P/&N')
    w.sheets[sheet_name].fit_to_pages(1, 1)


def bring_sheets_to_front(w, order):
    def place_first(x): return order.index(x.name) if x.name in order else len(order)
    w.book.worksheets_objs.sort(key = place_first)


def convert_named_ranges_to_print_areas(file):
    w = load_workbook(file)

    # openpyxl removes the axis from all charts for some reason
    # this happens when reading a workbook into memory and then saving that sames workbook
    for s in w.sheetnames:
        if len(w[s]._charts) > 0:
            for c in w[s]._charts:
                c.x_axis.delete = False
                c.y_axis.delete = False

    regions = [x for x in w.defined_names]
    sheets = [w.defined_names.get(x).value.split('!')[0] for x in regions]
    areas = [w.defined_names.get(x).value.split('!')[1] for x in regions]
    new_list = pd.Series(areas, index = sheets).groupby(level = 0).apply(lambda x: ', '.join(x))
    for i in range(len(new_list)):
        w[new_list.index[i]].print_area = new_list.values[i]
    w.save(file)


def add_table_of_contents(w):
    contents_sheet_name = "Contents"

    # get workbook name for use in hyperlink
    workbook_name = w.book.filename.name.split('\\')[-1]

    # get all the named regions.  this is the data used to create the table of contents
    regions = w.book.defined_names
    regions = [x for x in regions if 'Print_Area' not in x[0]]
    region_name = [x[0] for x in regions]
    sheet = [x[2].split('!')[0] for x in regions]

    # build the TOC dataset
    table_of_contents = pd.DataFrame({"sheet": sheet, "table_name": region_name})

    table_of_contents["table_name"] = table_of_contents.apply(
        lambda x: f'==HYPERLINK("[{workbook_name}]{x["sheet"]}!{x["table_name"]}", '
                  f''f'"{x["table_name"]}")',
        axis = 1
        )
    table_of_contents["item_number"] = range(len(table_of_contents))
    table_of_contents = table_of_contents[["item_number", "sheet", "table_name"]]
    table_of_contents.columns = proper_case(table_of_contents.columns)

    # style the TOC
    styled_context = table_of_contents.style
    styled_context.apply_index(column_format_standard, axis = 1)
    styled_context.apply_index(column_format_header_1, axis = 1)
    styled_context.apply(alternate_color_rows, dd = styled_context.data)
    styled_context.apply(format_hyperlink, subset = "Table Name")

    # Add worksheet and dataset
    w.book.add_worksheet(contents_sheet_name)
    styled_context.to_excel(w, sheet_name = "Contents", index = False)

    # Add TOC named region
    cell_range = get_dataframe_cell_range(table_of_contents, 0, 0)
    add_named_region(w, contents_sheet_name, cell_range, contents_sheet_name)

    # format worksheet
    w.sheets["Contents"].set_column(0, 0, 15)
    w.sheets["Contents"].set_column(1, 1, 25)
    w.sheets["Contents"].set_column(2, 2, 50)
